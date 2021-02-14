import requests
import os
import itertools
import re
import openpyxl
import csv
import sys
import json
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd
from jinja2 import Environment, FileSystemLoader
import pdfkit
import pydf


class NhlStats:
    def __init__(self, team, season, folder='default', titled_folder=True, export_type='xlsx'):
        self.team = team
        self.season = season
        self.folder = folder
        self.titled_folder = titled_folder
        self.export_type = export_type
        self.index = None
        self.dict_index = {
            'Anaheim-Ducks': 1580,
            'Arizona-Coyotes': 72,
            'Boston-Bruins': 52,
            'Buffalo-Sabres': 53,
            'Calgary-Flames': 54,
            'Carolina-Hurricanes': 55,
            'Chicago-Blackhawks': 56,
            'Colorado-Avalanche': 57,
            'Columbus-Blue-Jackets': 58,
            'Dallas-Stars': 59,
            'Detroit-Red-Wings': 60,
            'Edmonton-Oilers': 61,
            'Florida-Panthers': 62,
            'Los-Angeles-Kings': 79,
            'Minnesota-Wild': 63,
            'Montreal-Canadiens': 64,
            'Nashville-Predators': 65,
            'New-Jersey-Devils': 66,
            'New-York-Islanders': 67,
            'New-York-Rangers': 68,
            'Ottawa-Senators': 69,
            'Philadelphia-Flyers': 70,
            'Pittsburgh-Penguins': 71,
            'San-Jose-Sharks': 73,
            'St-Louis-Blues': 74,
            'Tampa-Bay-Lightning': 75,
            'Toronto-Maple-Leafs': 76,
            'Vancouver-Canucks': 77,
            'Vegas-Golden-Knights': 22211,
            'Washington-Capitals': 78,
            'Winnipeg-Jets': 9966
        }

        try:
            self.index = self.dict_index[team.title()]
        except KeyError:
            name_regex = re.compile(f'({team.lower()})')
            for city_name in self.dict_index.keys():
                try:
                    if name_regex.search(city_name.lower()):
                        self.index = self.dict_index[city_name]
                        team = city_name.lower()
                        self.team = city_name.title()
                        break
                except AttributeError:
                    continue

        if not self.index:
            print('Please write the full team name with - for spaces or part of it')
            sys.exit()

    def changed_name(self, team, season):
        lst_changed_names = [
            ["Arizona-Coyotes", "2014-2015", "Phoenix-Coyotes"],
            ["Anaheim-Ducks", "2006-2007", "Mighty-Ducks-Of-Anaheim"]
        ]
        for (new_name, first_used, old_name) in lst_changed_names:
            if team == new_name and season < first_used:
                return old_name

        return team

    def download(self):
        def get_soup():
            soup = BeautifulSoup(res.text, 'html.parser')
            base_css = 'table.table-striped.table-sortable.skater-stats.highlight-stats tbody + tbody tr:not(.space) '
            return ('A', soup.select(base_css + '.txt-blue')),\
                ('B', soup.select(base_css + 'td[class="gp"]')),\
                ('C', soup.select(base_css + 'td[class="g"]')),\
                ('D', soup.select(base_css + 'td[class="a"]')),\
                ('E', soup.select(base_css + 'td[class="tp sorted"]')),\
                ('F', soup.select(base_css + 'td[class="pim"]')),\
                ('G', soup.select(base_css + 'td[class="pm"]'))

        def get_data(data_type, index):
            letter = data_type[0]
            cell = f'{letter}{index + 2}'
            stats_sheet[cell].alignment = centered
            stats_sheet[cell].font = data_font
            if index % 2 == 0:
                stats_sheet[cell].fill = grey_fill
            data = data_type[1][index].getText().strip()
            stats_sheet[cell] = data

        def xlsx_setup():
            stats_wb = openpyxl.Workbook()
            stats_sheet = stats_wb['Sheet']
            stats_sheet.freeze_panes = 'B2'
            stats_sheet['A1'] = 'Player'
            stats_sheet['B1'] = 'GP'
            stats_sheet['C1'] = 'G'
            stats_sheet['D1'] = 'A'
            stats_sheet['E1'] = 'P'
            stats_sheet['F1'] = 'PIM'
            stats_sheet['G1'] = '+/-'
            for i in range(1, 8):
                letter = get_column_letter(i)
                stats_sheet[f'{letter}1'].alignment = centered

            return stats_wb, stats_sheet

        def export_xlsx(index):
            get_data(player_name, index)
            get_data(games_played, index)
            get_data(goals, index)
            get_data(assists, index)
            get_data(points, index)
            get_data(penalty_minutes, index)
            get_data(plus_minus, index)

        def cell_sizes():
            stats_sheet.column_dimensions['A'].width = 35
            for i in range(2, len(player_name[1]) + 2):
                stats_sheet.row_dimensions[i].height = 25

        def get_path(export_type):
            if self.folder == "default":
                if self.titled_folder:
                    filepath = Path(Path.cwd())/self.team.title() / \
                        f'{self.team.title()}_{self.season}.{export_type}'
                    os.makedirs(self.team.title(), exist_ok=True)
                else:
                    filepath = Path(Path.cwd())/self.team.lower() / \
                        f'{self.team.lower()}_{self.season}.{export_type}'
                    os.makedirs(self.team.lower(), exist_ok=True)

            else:
                if self.titled_folder:
                    folder = Path(self.folder)/self.team.title()
                    filepath = folder / \
                        f'{self.team.lower()}_{self.season}.{export_type}'
                else:
                    folder = Path(self.folder)/self.team.lower()
                    filepath = folder / \
                        f'{self.team.lower()}_{self.season}.{export_type}'
                os.makedirs(folder, exist_ok=True)
            return filepath

        team = self.changed_name(self.team, self.season)

        url = f'https://www.eliteprospects.com/team/{self.index}/{team.lower()}/{self.season}?tab=stats'
        res = requests.get(url)
        res.raise_for_status()

        centered = Alignment(horizontal='center')
        data_font = Font(size=14)
        grey_fill = PatternFill("solid", fgColor='00C0C0C0')

        stats_wb, stats_sheet = xlsx_setup()

        player_name, games_played, goals, assists, points, \
            penalty_minutes, plus_minus = get_soup()

        cell_sizes()

        for index in itertools.count():
            try:
                player_name[1][index]
                export_xlsx(index)

            except IndexError:
                print(f'Downloaded {self.team} {self.season}')
                break

        filepath = get_path(self.export_type)

        if self.export_type == "xlsx":
            stats_wb.save(filepath)

        elif self.export_type == "pdf-html":
            stats_wb.save(get_path("xlsx"))
            pd_df = pd.read_excel(str(get_path("xlsx")))

            env = Environment(loader=FileSystemLoader('.'))
            template = env.get_template('templates/download_stats.html')
            template_vars = {'stats_table': pd_df.to_html(
                justify='center', index=False, border=0, bold_rows=False, escape=False)}
            html_out = template.render(template_vars)

            html_out = re.sub(
                r'<table border="0" class="dataframe">',
                r'<table border="0" class="dataframe" cellspacing="0">',
                html_out
            )

            try:
                pdf = pydf.generate_pdf(html_out)
                with open(str(get_path("html")), 'wb') as pdf_file:
                    pdf_file.write(pdf)
            except:
                html_file = open(str(get_path("html")), 'w', encoding='utf8')
                html_file.write(html_out)
                html_file.close()

                config = pdfkit.configuration(
                    wkhtmltopdf='C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')

                pdfkit.from_file(str(get_path("html")), str(get_path(
                    "pdf")), configuration=config, css=str(Path.cwd()/'static'/'css'/'download_stats.css'), options={'quiet': ''})

        elif self.export_type == "csv":
            with open(filepath, 'w') as csv_file:
                csv_writer = csv.writer(csv_file)
                for row in stats_sheet.rows:
                    csv_writer.writerow([cell.value for cell in row])

        elif self.export_type == "json":
            json_file = {self.team: {self.season: {}}}
            for row in range(2, stats_sheet.max_row+1):
                json_file[self.team][self.season][stats_sheet[f'A{row}'].value] = {
                    "GP": stats_sheet[f'B{row}'].value,
                    "G": stats_sheet[f'C{row}'].value,
                    "A": stats_sheet[f'D{row}'].value,
                    "P": stats_sheet[f'E{row}'].value,
                    "PIM": stats_sheet[f'F{row}'].value,
                    "+/-": stats_sheet[f'g{row}'].value
                }
            with open(filepath, 'w') as export_file:
                out = json.dumps(json_file, indent=4)
                export_file.write(out)

        return str(filepath)
